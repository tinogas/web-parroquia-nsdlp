# -*- coding: utf-8 -*-
"""
extraer_agenda.py — Convierte «agenda 2026.xlsx» en una hoja revisable.

La agenda parroquial impresa se transcribió a un xlsx cuya única hoja,
«AGENDA 2026», guarda las actividades en CUADROS DE TEXTO FLOTANTES: las celdas
sólo tienen los números de día. Este script lee la geometría de esos cuadros,
la cruza con la rejilla del calendario y saca una fila por actividad con su
fecha real.

Cómo está montada la hoja (verificado contra el calendario de 2026):

  * Cada bloque de 8 columnas + 1 separadora trae DOS medios meses, así se
    imprime la agenda: DOM-MIE con fondo blanco de un mes, y JUE-SAB con fondo
    amarillo (FFFFFF00) de OTRO mes —el «mes mocho», que se completa con las
    columnas DOM-MIE de otro bloque—.
  * El nombre del mes blanco es un cuadro flotante encima de DOMINGO, al lado
    del de «Parroquia Nuestra Señora de la Paz». Los valores de las celdas de
    las filas 2 y 10 (A2='Junio', J2='Julio'…) son residuo de una transcripción
    anterior equivocada: se ignoran.
  * El mes de la mitad amarilla es el complemento a 12 del blanco: Junio<->Junio,
    Mayo<->Julio, Abril<->Agosto, Marzo<->Septiembre, Febrero<->Octubre,
    Enero<->Noviembre. Diciembre queda solo en el último bloque.

No se usa openpyxl para LEER: load_workbook() se cae con este archivo porque un
pitchFamily del drawing está fuera del rango que acepta. Se parsea el XML del
xlsx a mano. Sí se usa openpyxl para ESCRIBIR el resultado.

Uso:
    python herramientas/extraer_agenda.py ["agenda 2026.xlsx"] [-o salida.xlsx]
"""
from __future__ import annotations

import argparse
import calendar
import datetime
import io
import os
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict

SS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
XDR = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
DML = '{http://schemas.openxmlformats.org/drawingml/2006/main}'

ANIO = 2026
AMARILLO = 'FFFFFF00'
# Un cuadro se considera "sobre" una celda si la cubre al menos este tanto por
# uno de lo que cubre la celda que más solapa. Los cuadros de texto se dibujan
# más anchos que la celda y desbordan a las vecinas.
UMBRAL_SOLAPE = 0.30

MESES = {'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
         'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11,
         'diciembre': 12}
NOMMES = {v: k.capitalize() for k, v in MESES.items()}
# 0 = domingo … 6 = sábado, como la columna dia_semana de la tabla horarios.
DIAS = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
# weekday() de python: 0 = lunes … 6 = domingo
WD_PY = {'DOM': 6, 'LUN': 0, 'MAR': 1, 'MIE': 2, 'JUE': 3, 'VIE': 4, 'SAB': 5}

BLOQUES = [1, 10, 19, 28, 37, 46, 55]  # primera columna de cada bloque

# Centros tal como están en la tabla `centros` de parroquia_nsdlp, con las
# variantes que usa la agenda: «C. Jesús el Sr.», «Centro Jesús el Señor»,
# «C.San Pío», «SAN PIO»… El orden importa: las casas de oración van primero
# porque son más específicas que el centro que las agrupa.
# (patrón, nombre del centro, centros.id o None si no está en la tabla)
CENTROS = [
    (r'(?:casa\s+de\s+oraci[oó]n\s+)?san\s+lucas', 'Casa de Oración San Lucas', None),
    (r'(?:casa\s+de\s+oraci[oó]n\s+)?santo\s+tom[aá]s', 'Casa de Oración Santo Tomás', None),
    (r'(?:casa\s+de\s+oraci[oó]n\s+)?san\s+jos[eé]', 'Casa de Oración San José', None),
    (r'(?:c\.?\s*|centro\s+)?jes[uú]s\s+el\s+(?:se[nñ]or|sr\.?)', 'Jesús el Señor', 3),
    (r'(?:c\.?\s*|centro\s+)?san\s+p[ií]o(?:\s+de\s+pietrelcina)?',
     'San Pío de Pietrelcina', 2),
    (r'sede', 'Parroquia Nuestra Señora de la Paz', 1),
    (r'didec', 'Didec', None),
]
# Pastorales tal como están en la tabla `pastorales`.
PASTORALES = [
    (16, 'MESC', ['mesc', 'ministro extraordinario', 'ministros']),
    (18, 'Catecismo', ['catequesis', 'catequista', 'catecismo', 'prebautismal']),
    (21, 'Lectores', ['proclamador', 'lector']),
    (22, 'AMA', ['club ama', 'grupo ama', 'adulto mayor', 'abuelos']),
    (23, 'Raíces', ['raices', 'raíces']),
    (24, 'Grupo JECSA', ['jecsa']),
]


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #
def col_nombre(i: int) -> str:
    s = ''
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def sin_acentos(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def clave(s: str) -> str:
    """Texto normalizado para comparar: sin acentos, minúsculas, un solo espacio."""
    return re.sub(r'[^a-z0-9 ]', ' ', re.sub(r'\s+', ' ', sin_acentos(s).lower())).strip()


def desespaciar(texto: str) -> tuple[str, bool]:
    """«S E M A N A   D E   C A T E Q U E S I S» → «SEMANA DE CATEQUESIS».

    Las barras de periodo llevan el título con las letras separadas para que
    ocupe todo el ancho. Devuelve (texto, se_recompuso).
    """
    if '\n' in texto:
        partes = [desespaciar(l) for l in texto.split('\n')]
        return '\n'.join(p for p, _ in partes), any(f for _, f in partes)
    tokens = re.findall(r'\S+|\s+', texto.strip())
    letras = [t for t in tokens if not t.isspace()]
    if len(letras) < 4 or sum(1 for t in letras if len(t) == 1) / len(letras) < 0.7:
        return texto, False
    seps = [len(t) for t in tokens if t.isspace()]
    if not seps:
        return texto, False
    umbral = (min(seps) + max(seps)) / 2
    if min(seps) == max(seps):          # separación uniforme: todo es una palabra
        umbral = max(seps) + 1
    fuera = []
    for t in tokens:
        if t.isspace():
            if len(t) > umbral:
                fuera.append(' ')
        else:
            fuera.append(t)
    return re.sub(r'\s+', ' ', ''.join(fuera)).strip(), True


# --------------------------------------------------------------------------- #
# lectura del xlsx
# --------------------------------------------------------------------------- #
class Hoja:
    """Rejilla y cuadros de texto de la hoja AGENDA 2026."""

    def __init__(self, ruta: str):
        with zipfile.ZipFile(ruta) as z:
            self.styles = ET.fromstring(z.read('xl/styles.xml'))
            self.sheet = ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
            self.drawing = ET.fromstring(z.read('xl/drawings/drawing1.xml'))
            try:
                shared = ET.fromstring(z.read('xl/sharedStrings.xml'))
            except KeyError:
                shared = []
        self.cadenas = [''.join(t.text or '' for t in si.iter(SS + 't')) for si in shared]
        self._leer_estilos()
        self._leer_celdas()
        self._leer_dimensiones()
        self.cuadros = list(self._leer_cuadros())

    # -- estilos ------------------------------------------------------------ #
    def _leer_estilos(self):
        rellenos = []
        for f in self.styles.find(SS + 'fills'):
            pf = f.find(SS + 'patternFill')
            color = None
            if pf is not None and pf.get('patternType') not in (None, 'none'):
                fg = pf.find(SS + 'fgColor')
                if fg is not None:
                    color = fg.get('rgb') or ('tema' + str(fg.get('theme')))
            rellenos.append(color)
        self._xf = [int(x.get('fillId') or 0) for x in self.styles.find(SS + 'cellXfs')]
        self._rellenos = rellenos

    def _leer_celdas(self):
        self.valor, self.relleno = {}, {}
        for fila in self.sheet.find(SS + 'sheetData'):
            for c in fila:
                ref, s, t = c.get('r'), int(c.get('s') or 0), c.get('t')
                v = c.find(SS + 'v')
                x = v.text if v is not None else None
                if t == 's' and x is not None:
                    x = self.cadenas[int(x)]
                self.valor[ref] = x or ''
                self.relleno[ref] = self._rellenos[self._xf[s]] if s < len(self._xf) else None

    def _leer_dimensiones(self):
        # Ancho de columna: unidades de carácter -> píxeles -> EMU (1 px = 9525).
        def a_px(w):
            return int(((256 * w + int(128 / 7)) / 256) * 7)

        self.ancho = {}
        for c in self.sheet.find(SS + 'cols'):
            emu = a_px(float(c.get('width'))) * 9525
            for i in range(int(c.get('min')), int(c.get('max')) + 1):
                self.ancho[i] = emu
        # Alto de fila: puntos -> EMU (1 pt = 12700).
        self.alto = {int(r.get('r')): int(float(r.get('ht') or 14.4) * 12700)
                     for r in self.sheet.find(SS + 'sheetData')}
        self.max_col = max(self.ancho) if self.ancho else 0
        self.max_fila = max(self.alto) if self.alto else 0

    # -- cuadros de texto --------------------------------------------------- #
    @staticmethod
    def _texto(sp):
        tb = sp.find(XDR + 'txBody')
        if tb is None:
            return ''
        parrafos = [''.join(t.text or '' for t in p.iter(DML + 't'))
                    for p in tb.findall(DML + 'p')]
        while parrafos and not parrafos[0].strip():
            parrafos.pop(0)
        while parrafos and not parrafos[-1].strip():
            parrafos.pop()
        return '\n'.join(parrafos)

    @staticmethod
    def _firma_relleno(sp):
        """Firma del color de fondo, para reconocer los trozos de una misma barra.

        sysClr windowText/window son el mismo gris que schemeClr tx1/bg1, así que
        se normalizan; si no, los trozos de una barra no se reconocerían entre sí.
        """
        spPr = sp.find(XDR + 'spPr')
        if spPr is None:
            return None
        if spPr.find(DML + 'noFill') is not None:
            return None
        sf = spPr.find(DML + 'solidFill')
        if sf is None:
            return 'otro'
        srgb = sf.find(DML + 'srgbClr')
        if srgb is not None:
            return '#' + srgb.get('val')
        for tag, equis in ((DML + 'schemeClr', {}),
                           (DML + 'sysClr', {'windowText': 'tx1', 'window': 'bg1'})):
            e = sf.find(tag)
            if e is None:
                continue
            base = e.get('val')
            base = equis.get(base, base)
            lm = e.find(DML + 'lumMod')
            lo = e.find(DML + 'lumOff')
            return (base
                    + (f'/lm{lm.get("val")}' if lm is not None else '')
                    + (f'/lo{lo.get("val")}' if lo is not None else ''))
        return 'otro'

    def _solape(self, inicio, desfase, tamano, medidas, maximo):
        """[(índice, emu solapado)] de las celdas que cubre el intervalo dado."""
        x0, x1 = desfase, desfase + max(tamano, 1)
        pos, i = 0, inicio
        while i > 1 and pos > x0:          # desfase negativo: retrocede
            i -= 1
            pos -= medidas.get(i, 0)
        salida = []
        while pos < x1 and i <= maximo:
            w = medidas.get(i, 0)
            a, b = max(x0, pos), min(x1, pos + w)
            if b > a:
                salida.append((i, b - a))
            pos += w
            i += 1
        return salida

    def _leer_cuadros(self):
        for ancla in self.drawing:
            sp = ancla.find(XDR + 'sp')
            if sp is None:
                continue
            nombre = sp.find(XDR + 'nvSpPr').find(XDR + 'cNvPr').get('name', '')
            texto = self._texto(sp)
            es_barra = nombre.startswith('Rect')
            # Los cuadros de texto vacíos son decorativos; los rectángulos vacíos
            # son la continuación de una barra de periodo y sí importan.
            if not texto.strip() and not es_barra:
                continue
            d = ancla.find(XDR + 'from')
            fc = int(d.find(XDR + 'col').text) + 1
            fco = int(d.find(XDR + 'colOff').text)
            fr = int(d.find(XDR + 'row').text) + 1
            fro = int(d.find(XDR + 'rowOff').text)
            if ancla.tag.endswith('twoCellAnchor'):
                h = ancla.find(XDR + 'to')
                tc = int(h.find(XDR + 'col').text) + 1
                tco = int(h.find(XDR + 'colOff').text)
                tr = int(h.find(XDR + 'row').text) + 1
                tro = int(h.find(XDR + 'rowOff').text)
                cx = sum(self.ancho.get(c, 0) for c in range(fc, tc)) - fco + tco
                cy = sum(self.alto.get(r, 0) for r in range(fr, tr)) - fro + tro
            else:
                ext = ancla.find(XDR + 'ext')
                cx, cy = int(ext.get('cx')), int(ext.get('cy'))
            yield dict(
                cols=self._solape(fc, fco, cx, self.ancho, self.max_col),
                filas=self._solape(fr, fro, cy, self.alto, self.max_fila),
                relleno=self._firma_relleno(sp), barra=es_barra, texto=texto,
            )


# --------------------------------------------------------------------------- #
# mapa del calendario
# --------------------------------------------------------------------------- #
class Calendario:
    """Traduce (columna, fila) de la hoja a una fecha de 2026."""

    def __init__(self, hoja: Hoja):
        self.hoja = hoja
        self.avisos: list[str] = []
        self.dommie: dict[tuple[int, int], int] = {}
        self.colmes: dict[tuple[int, int], tuple[int, str]] = {}
        self.fecha: dict[tuple[int, int], datetime.date] = {}
        self._rotulos()
        self._columnas()
        self._fechas()

    @staticmethod
    def _bloque(col):
        for b in BLOQUES:
            if b <= col <= b + 7:
                return b
        for b in BLOQUES:                 # rótulo sobre la columna separadora
            if col == b - 1:
                return b
        return None

    def _rotulos(self):
        """Mes de cada mitad DOM-MIE, leído de los cuadros flotantes."""
        for c in self.hoja.cuadros:
            nombre = clave(c['texto'])
            if nombre not in MESES or not c['filas'] or not c['cols']:
                continue
            fila = c['filas'][0][0]
            if fila not in (1, 2, 9, 10):
                continue
            banda = 1 if fila <= 2 else 2
            bloque = self._bloque(c['cols'][0][0])
            if bloque:
                self.dommie[(banda, bloque)] = MESES[nombre]

    def _dias_del_mes(self, mes, wd_py):
        return {d for d in range(1, calendar.monthrange(ANIO, mes)[1] + 1)
                if datetime.date(ANIO, mes, d).weekday() == wd_py}

    def _columnas(self):
        for banda, (f_dias, f0, f1) in [(1, (3, 4, 8)), (2, (11, 12, 16))]:
            for b in BLOQUES:
                blanco = self.dommie.get((banda, b))
                for i, dia in [(0, 'DOM'), (1, 'LUN'), (2, 'MAR'), (3, 'MIE')]:
                    if blanco and self.hoja.valor.get(f'{col_nombre(b + i)}{f_dias}'):
                        self.colmes[(banda, b + i)] = (blanco, dia)
                amarilla = any(self.hoja.relleno.get(f'{col_nombre(b + i)}{f_dias}') == AMARILLO
                               for i in (5, 6, 7))
                if not amarilla:
                    for i, dia in [(5, 'JUE'), (6, 'VIE'), (7, 'SAB')]:
                        if blanco and self.hoja.valor.get(f'{col_nombre(b + i)}{f_dias}'):
                            self.colmes[(banda, b + i)] = (blanco, dia)
                    continue
                # Mitad amarilla: se deduce del patrón de números de día y se
                # confirma con la regla «mes amarillo = 12 - mes blanco».
                posibles = None
                for i, dia in [(5, 'JUE'), (6, 'VIE'), (7, 'SAB')]:
                    col = col_nombre(b + i)
                    if not self.hoja.valor.get(f'{col}{f_dias}'):
                        continue
                    vistos = {int(self.hoja.valor[f'{col}{r}'])
                              for r in range(f0, f1 + 1)
                              if self.hoja.valor.get(f'{col}{r}', '').strip().isdigit()}
                    cand = {m for m in range(1, 13)
                            if vistos and vistos <= self._dias_del_mes(m, WD_PY[dia])}
                    posibles = cand if posibles is None else posibles & cand
                if not posibles:
                    continue
                regla = (12 - blanco) if blanco else 12
                if regla in posibles:
                    mes = regla
                elif len(posibles) == 1:
                    mes = posibles.copy().pop()
                else:
                    self.avisos.append(
                        f'Bloque {col_nombre(b)} banda {banda}: mitad amarilla sin '
                        f'resolver, posibles {sorted(NOMMES[m] for m in posibles)}')
                    continue
                for i, dia in [(5, 'JUE'), (6, 'VIE'), (7, 'SAB')]:
                    if self.hoja.valor.get(f'{col_nombre(b + i)}{f_dias}'):
                        self.colmes[(banda, b + i)] = (mes, dia)

    def _fechas(self):
        for banda, (f0, f1) in [(1, (4, 8)), (2, (12, 16))]:
            for (bb, col), (mes, dia) in self.colmes.items():
                if bb != banda:
                    continue
                for r in range(f0, f1 + 1):
                    bruto = self.hoja.valor.get(f'{col_nombre(col)}{r}', '').strip()
                    if not bruto.isdigit():
                        continue
                    ref = f'{col_nombre(col)}{r}'
                    try:
                        f = datetime.date(ANIO, mes, int(bruto))
                    except ValueError:
                        self.avisos.append(
                            f'{ref}: «{bruto}» no es un día válido de {NOMMES[mes]}')
                        continue
                    if f.weekday() != WD_PY[dia]:
                        self.avisos.append(
                            f'{ref}: el {bruto} de {NOMMES[mes]} no cae en {dia}; '
                            f'número mal transcrito en la hoja')
                        continue
                    self.fecha[(col, r)] = f

    # -- consulta ----------------------------------------------------------- #
    def celdas(self, cuadro, umbral=UMBRAL_SOLAPE):
        cols = self._principales(cuadro['cols'], umbral)
        filas = self._principales(cuadro['filas'], umbral)
        return [(c, r) for c in cols for r in filas]

    @staticmethod
    def _principales(lista, umbral):
        if not lista:
            return []
        tope = max(s for _, s in lista)
        return [i for i, s in lista if s >= umbral * tope]

    def fechas_de(self, cuadro, umbral=UMBRAL_SOLAPE):
        f = sorted({self.fecha[cr] for cr in self.celdas(cuadro, umbral) if cr in self.fecha})
        if not f:   # sin umbral, para cuadros muy desplazados
            f = sorted({self.fecha[(c, r)] for c, _ in cuadro['cols']
                        for r, _ in cuadro['filas'] if (c, r) in self.fecha})
        return f

    def ancla(self, cuadro):
        col = max(cuadro['cols'], key=lambda x: x[1])[0] if cuadro['cols'] else None
        fila = max(cuadro['filas'], key=lambda x: x[1])[0] if cuadro['filas'] else None
        return f'{col_nombre(col) if col else "?"}{fila if fila else "?"}'


# --------------------------------------------------------------------------- #
# interpretación del texto
# --------------------------------------------------------------------------- #
RE_RANGO = re.compile(
    r'(\d{1,2})(?::(\d{2}))?\s*(a\.?m\.?|p\.?m\.?)?\s*(?:-|a|hasta)\s*'
    r'(\d{1,2})(?::(\d{2}))?\s*(a\.?m\.?|p\.?m\.?)', re.IGNORECASE)
RE_SUELTA = re.compile(r'(\d{1,2})(?::(\d{2}))?\s*(a\.?m\.?|p\.?m\.?)', re.IGNORECASE)


def _a_24h(h, m, sufijo):
    h, m = int(h), int(m or 0)
    if sufijo:
        s = sufijo.lower().replace('.', '')
        # h > 12 ya viene en 24 h («19:45-21:15 pm» en una copia del panel).
        if s == 'pm' and h < 12:
            h += 12
        elif s == 'am' and h == 12:
            h = 0
    return h, m


def _en_parentesis(texto: str, pos: int) -> bool:
    """¿La posición cae dentro de un paréntesis? «Ensayo (Misa de 5 pm)»: el 5 pm
    es una aclaración del título, no la hora de la actividad."""
    return texto.count('(', 0, pos) > texto.count(')', 0, pos)


# Cadena de horas de la misma actividad: «5 pm y 7 pm», «9 am, 12 y 6 pm»,
# «8 y 10 am y 12 pm». El separador es una coma, una «y»/«e», o las dos.
# El «\s*» del sufijo va DENTRO del grupo opcional: si queda fuera se come el
# espacio que necesita el separador siguiente y la cadena se corta a la mitad
# («9 am, 12 y 6 pm» se leería como «9 am, 12»).
RE_CADENA = re.compile(
    r'(\d{1,2})(?::(\d{2}))?(?:\s*(a\.?m\.?|p\.?m\.?))?'
    r'(?:(?:\s*,\s*|\s+(?:y|e)\s+|\s*,\s*(?:y|e)\s+)'
    r'(\d{1,2})(?::(\d{2}))?(?:\s*(a\.?m\.?|p\.?m\.?))?)+', re.IGNORECASE)
RE_UNA = re.compile(r'(\d{1,2})(?::(\d{2}))?(?:\s*(a\.?m\.?|p\.?m\.?))?', re.IGNORECASE)


def cadena_horas(texto: str):
    """(horas, inicio, fin) de una cadena de horas de la misma actividad.

    Devuelve (None, None, None) si el texto no trae una. El span sirve para
    quitar la cadena completa del título: si sólo se recorta la primera hora,
    quedan restos como «Misa Comunión 8».
    """
    m = RE_CADENA.search(texto)
    if not m or _en_parentesis(texto, m.start()) or RE_RANGO.search(texto):
        return None, None, None
    encontradas = [s for s in RE_UNA.finditer(m.group(0)) if s.group(1)]
    if len(encontradas) < 2:
        return None, None, None
    ultimo = next((s.group(3) for s in reversed(encontradas) if s.group(3)), None)
    if not ultimo:
        return None, None, None
    # Cada una con su sufijo; si falta, hereda el de la última de la cadena.
    horas = [_a_24h(s.group(1), s.group(2), s.group(3) or ultimo) for s in encontradas]
    # Las horas de una cadena van de menor a mayor, así que las de antes del
    # mediodía no deben heredar el p.m. final: «8 y 10 am y 12 pm» es 8, 10 y 12,
    # no 20, 10 y 12. Se corrige de derecha a izquierda.
    for i in range(len(horas) - 2, -1, -1):
        if horas[i] >= horas[i + 1] and not encontradas[i].group(3):
            alt = _a_24h(encontradas[i].group(1), encontradas[i].group(2), 'am')
            if alt < horas[i + 1]:
                horas[i] = alt
    return [f'{h:02d}:{m_:02d}' for h, m_ in horas], m.start(), m.end()


def horas_alternativas(texto: str):
    """La misma actividad a varias horas: «Misa 5 pm y 7 pm» son dos misas."""
    return cadena_horas(texto)[0] or []


def parsear_horas(texto: str):
    """(hora, hora_fin, resto_del_texto, dudas). Horas como 'HH:MM' o None."""
    dudas = []
    m = RE_RANGO.search(texto)
    if m and _en_parentesis(texto, m.start()):
        m = next((x for x in RE_RANGO.finditer(texto)
                  if not _en_parentesis(texto, x.start())), None)
    if m:
        h1, m1, s1, h2, m2, s2 = m.groups()
        fin = _a_24h(h2, m2, s2)
        ini = _a_24h(h1, m1, s1 or s2)      # «6 a 8 pm»: el 6 hereda el pm
        if ini > fin:
            # «10-12pm» es de 10 de la mañana a mediodía, no de las 22 h.
            alt = _a_24h(h1, m1, 'am')
            if alt < fin:
                ini = alt
                dudas.append('rango de horas corregido: se leyó la de inicio como a.m.')
            else:
                dudas.append('rango de horas incoherente')
        resto = texto[:m.start()] + ' ' + texto[m.end():]
        return f'{ini[0]:02d}:{ini[1]:02d}', f'{fin[0]:02d}:{fin[1]:02d}', resto, dudas

    # Cadena de horas de la misma actividad: se recorta entera del título, o
    # quedan restos como «Misa Comunión 8» o «Misas de Confirmación y 7 pm».
    encadenadas, ini_c, fin_c = cadena_horas(texto)
    if encadenadas:
        return encadenadas[0], None, texto[:ini_c] + ' ' + texto[fin_c:], dudas

    sueltas = [s for s in RE_SUELTA.finditer(texto) if not _en_parentesis(texto, s.start())]
    if sueltas:
        if len(sueltas) > 1:
            dudas.append('varias horas en la misma línea: '
                         + ', '.join(s.group(0).strip() for s in sueltas))
        s = sueltas[0]
        h, mi = _a_24h(s.group(1), s.group(2), s.group(3))
        return f'{h:02d}:{mi:02d}', None, texto[:s.start()] + ' ' + texto[s.end():], dudas

    # Hora sin am/pm: «Misa 10:30 Jesús el Señor».
    m = re.search(r'(?<![\d:])(\d{1,2}):(\d{2})(?![\d:])', texto)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if h < 7:
            h += 12
        dudas.append(f'hora «{m.group(0)}» sin a.m./p.m.: se interpretó {h:02d}:{mi:02d}')
        resto = texto[:m.start()] + ' ' + texto[m.end():]
        return f'{h:02d}:{mi:02d}', None, resto, dudas
    return None, None, texto, dudas


RE_CENTROS = [(re.compile(rf'\b{p}\b', re.IGNORECASE), nombre, cid)
              for p, nombre, cid in CENTROS]
# Lo que se recorta del título: el nombre del centro y su preposición delante.
PALABRAS_LUGAR = re.compile(
    r'\b(?:en\s+)?(?:la\s+|el\s+)?(?:' + '|'.join(p for p, _, _ in CENTROS) + r')\b',
    re.IGNORECASE)


def detectar_lugar(texto: str):
    for rx, nombre, _cid in RE_CENTROS:
        if rx.search(texto):
            return nombre
    return None


def centro_id(nombre: str | None):
    for _rx, n, cid in RE_CENTROS:
        if n == nombre:
            return cid
    return None


def detectar_pastoral(texto: str):
    k = clave(texto)
    for pid, nombre, pistas in PASTORALES:
        if any(p in k for p in pistas):
            return pid, nombre
    return None, None


RESTOS = re.compile(r'^[\s,;:.\-–—()]+|[\s,;:.\-–—()]+$')
FUNCIONALES = re.compile(r'\s*\b(?:de|del|en|el|la|los|las|y|e|a|con|para)\s*$',
                         re.IGNORECASE)


def _vaciar_parentesis(texto: str) -> str:
    """Borra el paréntesis que ya no dice nada tras recortarle hora y lugar.

    «Inicia Novena de Navidad ( Sede / 6:00 pm )» → «Inicia Novena de Navidad»,
    pero «Ensayo de Confirmación (Misa de 5 pm)» se queda entero: ahí el
    paréntesis es una aclaración de verdad.
    """
    def uno(m):
        dentro = RE_SUELTA.sub(' ', m.group(1))
        dentro = PALABRAS_LUGAR.sub(' ', dentro)
        dentro = re.sub(r'\b[ap]\.?\s?m\.?\b', ' ', dentro, flags=re.IGNORECASE)
        letras = re.sub(r'[^^\w]|[\d_]', '', dentro, flags=re.UNICODE)
        return m.group(0) if len(letras) >= 3 else ' '
    return re.sub(r'\(([^)]*)\)', uno, texto)


def limpiar_titulo(texto: str) -> str:
    """Quita del texto la mención del lugar y los restos de puntuación.

    Las fiestas litúrgicas están escritas en mayúsculas en la agenda y se dejan
    tal cual: pasarlas a Título estropea las siglas (MESC, VII) y pone en
    mayúscula preposiciones que en español van en minúscula.
    """
    t = PALABRAS_LUGAR.sub(' ', texto)
    t = _vaciar_parentesis(t)
    t = re.sub(r'\s+', ' ', t).strip()
    # Al quitar el lugar suelen quedar dos o tres palabras funcionales colgando
    # («…Llama de Amor en la»), así que se recortan una tras otra.
    anterior = None
    while t != anterior:
        anterior = t
        t = RESTOS.sub('', FUNCIONALES.sub('', t))
    # Paréntesis que quedó abierto al recortar («Ensayo (Misa de 5 pm»).
    if t.count('(') > t.count(')'):
        t += ')'
    return t


def partir_por_horas(linea: str):
    """«Rosario Fátima 7pm Hora Santa 8pm» → dos actividades, una por hora.

    Distinto de horas_alternativas(), que es para la misma actividad repetida a
    varias horas («Misa 5 pm y 7 pm»): aquí entre una hora y la siguiente hay
    texto, y ese texto es el título de la segunda.
    """
    if RE_RANGO.search(linea) or horas_alternativas(linea):
        return [linea]
    marcas = [s for s in RE_SUELTA.finditer(linea) if not _en_parentesis(linea, s.start())]
    if len(marcas) < 2:
        return [linea]
    crudos, desde = [], 0
    for s in marcas:
        crudos.append(linea[desde:s.end()].strip(' ,;'))
        desde = s.end()
    cola = linea[desde:].strip(' ,;')
    if crudos and cola:      # el lugar suele ir al final: «… 8pm Sede»
        crudos[-1] += ' ' + cola
    return heredar_titulos(crudos) or [linea]


def heredar_titulos(trozos):
    """Un trozo sin título propio, sólo lugar y hora, es otra sesión de la misma
    actividad: «TALLER … Sede 6pm San Pío 4pm Jesús el Sr 5pm» son tres talleres,
    y «MISA COMUNION C. Jesús el Señor 10 am y 12 pm» / «Centro San Pío 7 y 8 pm»
    son la misma misa en dos centros."""
    salida, previo = [], None
    for crudo in trozos:
        propio = limpiar_titulo(RE_SUELTA.sub(' ', RE_CADENA.sub(' ', crudo)))
        if clave(propio):
            previo = propio
            salida.append(crudo)
        elif previo:
            salida.append(f'{previo} {crudo}')
        else:
            salida.append(crudo)
    return salida


def partir_actividades(texto: str):
    """Un cuadro trae varias actividades. Devuelve la lista de textos.

    Dos separadores: el párrafo vacío corta de verdad, y la línea que trae la
    hora cierra la actividad en curso («Misa Casa de Oración» + «San José 6 pm»
    son una sola). Tras este segundo corte, un trozo que no trae hora suele ser
    la continuación del anterior —el lugar en la línea de abajo, «Rosario y
    meditación», «y confesiones»— y se le vuelve a pegar.
    """
    bloques, actual, debil = [], [], []
    def cerrar(por_hora):
        nonlocal actual
        if actual:
            debil.append((' '.join(actual), por_hora))
            actual = []

    for linea in texto.split('\n'):
        s = re.sub(r'\s+', ' ', linea).strip(' \'"')
        if not s:
            cerrar(False)
            if debil:
                bloques.append(list(debil))
                debil.clear()
            continue
        actual.append(s)
        if RE_SUELTA.search(s) or RE_RANGO.search(s):
            cerrar(True)
    cerrar(False)
    if debil:
        bloques.append(list(debil))

    salida = []
    for bloque in bloques:
        trozos = []
        for texto_trozo, por_hora in bloque:
            tiene_hora = por_hora or bool(RE_SUELTA.search(texto_trozo))
            if trozos and not tiene_hora:
                trozos[-1] += ' ' + texto_trozo
            else:
                trozos.append(texto_trozo)
        partidos = []
        for t in trozos:
            partidos.extend(partir_por_horas(t))
        salida.extend(heredar_titulos(partidos))
    return [s for s in salida if len(clave(s)) > 2]


# --------------------------------------------------------------------------- #
# panel de actividades semanales
# --------------------------------------------------------------------------- #
def es_panel_semanal(texto: str) -> bool:
    k = clave(texto)
    return 'actividades' in k and 'semanales' in k


ORDEN_DIAS = ['domingo', 'lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
RE_DIA = re.compile(r'\b(lunes|martes|mi[eé]rcoles|jueves|viernes|s[aá]bados?|domingos?)\b',
                    re.IGNORECASE)


def _indice_dia(nombre: str):
    """«Miércoles» → 3, «SABADOS» → 6. 0 = domingo, como horarios.dia_semana.

    Ojo con la «s»: lunes, martes, miércoles, jueves y viernes ya la llevan; sólo
    «sábados» y «domingos» son plurales de verdad.
    """
    n = sin_acentos(nombre).strip().lower()
    if n not in ORDEN_DIAS and n.endswith('s') and n[:-1] in ORDEN_DIAS:
        n = n[:-1]
    return ORDEN_DIAS.index(n) if n in ORDEN_DIAS else None


# Los encabezados del panel van en mayúsculas y sin tildes por diseño, pero son
# el nombre de la actividad y acaban publicados en la página de horarios.
SECCIONES = {
    'santo rosario': 'Santo Rosario',
    'casas de oracion comunidad san pio': 'Casa de Oración (Comunidad San Pío)',
}


def nombre_de_seccion(seccion: str) -> str:
    conocida = SECCIONES.get(clave(seccion))
    if conocida:
        return conocida
    if not seccion.isupper():
        return seccion
    # Title case a la española: las palabras funcionales se quedan en minúscula.
    palabras = [p.title() if i == 0 or not re.fullmatch(ATADURA, p.lower()) else p.lower()
                for i, p in enumerate(seccion.split())]
    return ' '.join(palabras)


def parsear_panel(texto: str):
    """El panel lateral «ACTIVIDADES SEMANALES» → filas de recurrencia semanal.

    El panel se repite idéntico en cada bloque de la hoja: es UNA lista, no doce.
    Su forma es un encabezado de día en mayúsculas (MARTES, MIERCOLES…) y debajo
    las actividades de ese día; más dos secciones sin día propio —SANTO ROSARIO y
    CASAS DE ORACION— donde el día va escrito en la línea de cada actividad.
    """
    filas, dias_actuales, seccion, buffer_ = [], None, None, []
    arrastra = False

    def cerrar():
        nonlocal buffer_, arrastra
        arrastra = False
        if not buffer_:
            return
        crudo = re.sub(r'\s+', ' ', ' '.join(buffer_)).strip()
        buffer_ = []
        if not clave(crudo):
            return
        # «Lunes a viernes 6pm», «sábado y domingo 5 pm», «lunes San Lucas 5pm»:
        # en las secciones sin encabezado de día (SANTO ROSARIO, CASAS DE ORACION)
        # el día va escrito en la propia línea. Si hay encabezado activo manda el
        # encabezado: bajo JUEVES, «Jueves Eucarístico» es el nombre de la
        # actividad, no la etiqueta del día.
        propios = []
        if dias_actuales is None:
            rango = re.search(
                r'(lunes|martes|mi[eé]rcoles|jueves|viernes|s[aá]bado|domingo)s?'
                r'\s+a\s+(lunes|martes|mi[eé]rcoles|jueves|viernes|s[aá]bado|domingo)s?',
                crudo, re.IGNORECASE)
            if rango:
                i, j = _indice_dia(rango.group(1)), _indice_dia(rango.group(2))
                propios = [d % 7 for d in range(i, (j if j >= i else j + 7) + 1)]
            else:
                propios = [d for d in (_indice_dia(n) for n in RE_DIA.findall(crudo))
                           if d is not None]
        objetivo = propios or dias_actuales or [None]

        # El nombre del día sólo se borra del título si es la etiqueta del día;
        # en «Jueves Eucarístico» forma parte del nombre de la actividad.
        sin_dias = RE_DIA.sub(
            lambda m: '' if _indice_dia(m.group(0)) in propios else m.group(0), crudo)
        sin_dias = re.sub(r'^\s*(?:a|y|e)\s+', ' ', sin_dias)
        sin_dias = re.sub(r'\s+(?:a|y|e)\s+(?=\d|$)', ' ', sin_dias)
        lugar_linea = detectar_lugar(crudo)
        pid, pnom = detectar_pastoral(crudo)
        # «Rosario N. Sra. Fátima 7pm Hora Santa 8pm» son dos actividades distintas.
        for segmento in partir_por_horas(sin_dias):
            hora, hora_fin, resto, dudas = parsear_horas(segmento)
            titulo = limpiar_titulo(resto)
            if not clave(titulo) and seccion:
                titulo = nombre_de_seccion(seccion)
            if not clave(titulo):
                continue
            if not propios and not dias_actuales:
                dudas.append('sin día de la semana')
            # «Misa 5 pm y 7 pm»: la misma actividad a dos horas.
            for h in (horas_alternativas(segmento) or [hora]):
                for d in objetivo:
                    filas.append(dict(dia=d, hora=h, hora_fin=hora_fin, titulo=titulo,
                                      lugar=detectar_lugar(segmento) or lugar_linea,
                                      pastoral_id=pid, pastoral=pnom,
                                      original=crudo, dudas='; '.join(dudas)))

    for linea in texto.split('\n'):
        s = re.sub(r'\s+', ' ', linea).strip()
        k = clave(s)
        if not s:
            cerrar()
            continue
        if k in ('actividades', 'semanales', 'actividades semanales',
                 'actividades semanales todo el ano'):
            continue
        hay_hora = bool(RE_SUELTA.search(s) or RE_RANGO.search(s))
        # Encabezado de día suelto: MARTES, MIERCOLES, SABADOS…
        if not hay_hora and RE_DIA.fullmatch(s.strip()):
            cerrar()
            dias_actuales = [_indice_dia(s.strip())]
            seccion = None
            continue
        # Encabezado de sección en mayúsculas sin día: SANTO ROSARIO,
        # CASAS DE ORACION / COMUNIDAD SAN PIO (dos líneas).
        if not hay_hora and s.isupper():
            cerrar()
            seccion = f'{seccion} {s}'.strip() if seccion and not dias_actuales else s
            dias_actuales = None
            continue
        colgante = bool(re.search(r'\b(?:en|de|del|y|a)$', s, re.IGNORECASE))
        buffer_.append(s)
        # Una línea que acaba en preposición sigue en la de abajo («Santo Rosario
        # 9am en» + «Centro Jesús el Señor»): se aguanta esa línea más, pero sólo
        # una, o la actividad se comería las siguientes.
        if arrastra or (hay_hora and not colgante):
            cerrar()
        arrastra = hay_hora and colgante
    cerrar()

    vistos, unicas = set(), []
    for f in filas:
        k = (f['dia'], f['hora'], clave(f['titulo']))
        if k in vistos:
            continue
        vistos.add(k)
        unicas.append(f)
    unicas.sort(key=lambda f: (f['dia'] if f['dia'] is not None else 9, f['hora'] or '99'))
    return unicas


# --------------------------------------------------------------------------- #
# barras de periodo
# --------------------------------------------------------------------------- #
def _palabras_fuertes(texto: str) -> set:
    return {p for p in clave(texto).split() if len(p) > 4}


ATADURA = r'(?:de|del|en|el|la|los|las|y|e|a|al|con|para|por|sin)'


def _continua(anterior: str, siguiente: str) -> bool:
    """¿«siguiente» es la continuación del título de «anterior»?

    Muchas barras comparten color, así que el color y la contigüidad de fechas no
    bastan: sin esta comprobación se pegarían periodos distintos («Encuentro
    Matrimonios Misioneros» con «Reunión Consejo OMPE en Línea»). Se unen si:

      * uno de los dos no trae texto (rectángulo de continuación);
      * comparten alguna palabra con peso («ENCUENTRO DE VARONES» / «ENCUENTRO
        VARONES»);
      * el primero acaba en preposición o el segundo empieza por una («ASAMBLEA
        DE» / «DIRECTORES OMPE», «CIERRE DE CICLO» / «DE CATEQUESIS»);
      * el segundo es un rabo de una o dos palabras («SEMANA DE CATEQUESIS» /
        «PREMATRIMONIAL», «Escuela de Animación» / «Misionera (ESAM)»).
    """
    a, b = clave(anterior), clave(siguiente)
    if not a or not b:
        return True
    if _palabras_fuertes(anterior) & _palabras_fuertes(siguiente):
        return True
    if re.match(rf'{ATADURA}\b', b):
        return True
    # Que el primero acabe en preposición no basta si el segundo ya es un título
    # entero por su cuenta: «Novena a Nuestra Señora de» no continúa en «Reunión
    # Consejo OMPE en Línea», que son tres palabras con peso y ninguna compartida.
    if re.search(rf'\b{ATADURA}$', a) and len(_palabras_fuertes(siguiente)) < 3:
        return True
    return len(b.split()) <= 2


def unir_barras(barras):
    """Une los trozos de una misma barra: mismo color, fechas que se continúan y
    títulos que encajan.

    Cuando un periodo cruza de semana, la agenda lo dibuja como varios
    rectángulos con el título troceado («N o v e n a  a  N u e s t r a
    S e ñ o r a  d e» + «de la Paz»).
    """
    grupos = defaultdict(list)
    for b in barras:
        grupos[b['relleno']].append(b)
    salida = []
    for relleno, lote in grupos.items():
        lote.sort(key=lambda b: (b['fechas'][0], b['fechas'][-1]))
        actual = None
        for b in lote:
            texto = desespaciar(b['cuadro']['texto'])[0]
            if (actual and (b['fechas'][0] - actual['fechas'][-1]).days <= 1
                    and _continua(actual['texto'], texto)):
                actual['fechas'] = sorted(set(actual['fechas']) | set(b['fechas']))
                actual['trozos'].append(b)
                actual['texto'] = (actual['texto'] + ' ' + texto).strip()
                continue
            if actual:
                salida.append(actual)
            actual = dict(relleno=relleno, fechas=list(b['fechas']), trozos=[b],
                          texto=texto)
        if actual:
            salida.append(actual)
    return salida


# --------------------------------------------------------------------------- #
# proceso
# --------------------------------------------------------------------------- #
def extraer(ruta: str):
    hoja = Hoja(ruta)
    cal = Calendario(hoja)

    paneles, barras, cuadros = [], [], []
    for c in hoja.cuadros:
        if es_panel_semanal(c['texto']):
            paneles.append(c)
        elif clave(c['texto']) in MESES or clave(c['texto']).startswith('parroquia nuestra'):
            continue
        elif c['barra']:
            f = cal.fechas_de(c)
            if f:
                barras.append(dict(fechas=f, relleno=c['relleno'], cuadro=c,
                                   ancla=cal.ancla(c)))
            else:
                cuadros.append(c)     # sin fecha: se reporta como duda
        else:
            cuadros.append(c)

    eventos = []

    # -- barras de periodo -------------------------------------------------- #
    for grupo in unir_barras(barras):
        textos, anclas, recompuesto = [], [], False
        for t in grupo['trozos']:
            limpio, cambiado = desespaciar(t['cuadro']['texto'])
            recompuesto = recompuesto or cambiado
            limpio = re.sub(r'\s+', ' ', limpio).strip()
            if limpio and clave(limpio) not in {clave(x) for x in textos}:
                textos.append(limpio)
            anclas.append(t['ancla'])
        crudo = ' '.join(textos).strip()
        if not crudo:
            continue
        ini, fin = grupo['fechas'][0], grupo['fechas'][-1]
        comunes = []
        if (fin - ini).days + 1 != len(grupo['fechas']):
            comunes.append('los días de la barra no son consecutivos: '
                           + ', '.join(str(f) for f in grupo['fechas']))
        if len(grupo['trozos']) > 1:
            comunes.append(f'periodo armado con {len(grupo["trozos"])} trozos de barra')
        if recompuesto:
            comunes.append('título con las letras separadas, recompuesto')
        # «TALLER … Sede 6pm San Pío 4pm Jesús el Sr 5pm»: una fila por sede.
        for segmento in partir_por_horas(crudo):
            hora, hora_fin, resto, dudas = parsear_horas(segmento)
            pid, pnom = detectar_pastoral(segmento)
            eventos.append(dict(
                inicio=ini, fin=fin if fin != ini else None, hora=hora,
                hora_fin=hora_fin, titulo=limpiar_titulo(resto) or segmento,
                lugar=detectar_lugar(segmento), pastoral_id=pid, pastoral=pnom,
                clase='periodo', origen=' + '.join(anclas), original=segmento,
                dudas='; '.join(comunes + dudas)))

    # -- cuadros de un día -------------------------------------------------- #
    for c in cuadros:
        fechas = cal.fechas_de(c)
        base_dudas = []
        if not fechas:
            base_dudas.append('el cuadro no cae sobre ningún día de la rejilla')
            fecha = None
        else:
            fecha = fechas[0]
            if len(fechas) > 1:
                base_dudas.append('el cuadro toca varios días ('
                                  + ', '.join(str(f) for f in fechas)
                                  + '): se asignó al primero')
        limpio, recompuesto = desespaciar(c['texto'])
        for trozo in partir_actividades(limpio):
            hora, hora_fin, resto, dudas = parsear_horas(trozo)
            pid, pnom = detectar_pastoral(trozo)
            titulo = limpiar_titulo(resto)
            if not clave(titulo):
                titulo = re.sub(r'\s+', ' ', trozo).strip()
            if recompuesto:
                dudas = dudas + ['título con las letras separadas, recompuesto']
            # «Misas de Confirmación 5 pm y 7 pm» son dos celebraciones: una fila
            # por hora, que es lo que se pidió al elegir «un evento por actividad».
            varias = horas_alternativas(trozo)
            for h in (varias or [hora]):
                eventos.append(dict(
                    inicio=fecha, fin=None, hora=h, hora_fin=hora_fin if not varias else None,
                    titulo=titulo, lugar=detectar_lugar(trozo), pastoral_id=pid,
                    pastoral=pnom, clase='evento', origen=cal.ancla(c), original=trozo,
                    dudas='; '.join(base_dudas + dudas
                                    + ([f'la línea traía {len(varias)} horas '
                                        f'({", ".join(varias)}): se hizo una fila por cada una']
                                       if varias else []))))

    # -- marcar las mensuales recurrentes ----------------------------------- #
    repes = Counter(clave(e['titulo']) for e in eventos if e['clase'] == 'evento')
    for e in eventos:
        if e['clase'] == 'evento' and repes[clave(e['titulo'])] >= 6:
            e['clase'] = 'mensual'

    # -- panel semanal ------------------------------------------------------ #
    semanales = []
    if paneles:
        mejor = max(paneles, key=lambda c: len(c['texto']))
        semanales = parsear_panel(mejor['texto'])

    eventos.sort(key=lambda e: (e['inicio'] or datetime.date(2099, 1, 1),
                                e['hora'] or '99', e['titulo']))
    return hoja, cal, eventos, semanales, len(paneles)


# --------------------------------------------------------------------------- #
# salida
# --------------------------------------------------------------------------- #
def horarios_existentes():
    """Los horarios ya cargados, para no volver a meterlos. [] si no hay BD."""
    try:
        import subprocess
        mysql = r'C:\xampp\mysql\bin\mysql.exe'
        if not os.path.isfile(mysql):
            return None
        r = subprocess.run(
            [mysql, '-u', 'root', '--default-character-set=utf8mb4', '-B', '--skip-column-names',
             'parroquia_nsdlp', '-e',
             'SELECT h.dia_semana, TIME_FORMAT(h.hora, "%H:%i"), h.tipo, '
             'COALESCE(c.nombre, ""), COALESCE(h.nota, "") '
             'FROM horarios h LEFT JOIN centros c ON c.id = h.centro_id WHERE h.activo = 1'],
            capture_output=True, text=True, encoding='utf-8', timeout=20)
        if r.returncode != 0:
            return None
        # Una nota vacía puede venir sin su tabulador, así que se rellena a 5.
        return [(l.split('\t') + [''] * 5)[:5]
                for l in r.stdout.strip().splitlines() if l.strip()]
    except Exception:
        return None


def escribir(ruta_salida, hoja, cal, eventos, semanales, n_paneles):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    cab = Font(bold=True, color='FFFFFF')
    fondo = PatternFill('solid', fgColor='1E4D8B')
    arriba = Alignment(vertical='top', wrap_text=True)

    def encabezar(ws, columnas):
        ws.append(columnas)
        for i, _ in enumerate(columnas, 1):
            c = ws.cell(row=1, column=i)
            c.font, c.fill = cab, fondo
        ws.freeze_panes = 'A2'

    def anchos(ws, medidas):
        for i, w in enumerate(medidas, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for fila in ws.iter_rows(min_row=2):
            for c in fila:
                c.alignment = arriba

    # --- Eventos ---
    ws = wb.active
    ws.title = 'Eventos'
    COLS = ['Fecha inicio', 'Fecha fin', 'Hora', 'Hora fin', 'Todo el día', 'Título',
            'Lugar', 'Pastoral', 'Clasificación', 'Celda origen', 'Texto original', 'Dudas']
    encabezar(ws, COLS)
    for e in eventos:
        ws.append([e['inicio'].isoformat() if e['inicio'] else '',
                   e['fin'].isoformat() if e['fin'] else '',
                   e['hora'] or '', e['hora_fin'] or '',
                   'sí' if not e['hora'] else '', e['titulo'], e['lugar'] or '',
                   e['pastoral'] or '', e['clase'], e['origen'], e['original'],
                   e['dudas']])
    anchos(ws, [13, 12, 8, 9, 11, 46, 30, 14, 14, 14, 52, 46])

    # --- Horarios semanales ---
    ws = wb.create_sheet('Horarios semanales')
    encabezar(ws, ['Día', 'dia_semana', 'Hora', 'Hora fin', 'Actividad', 'Lugar',
                   'Pastoral', '¿Ya existe?', 'Texto original', 'Dudas'])
    existentes = horarios_existentes()
    for s in semanales:
        ya = ''
        if existentes is None:
            ya = 'sin comprobar (no se pudo consultar la BD)'
        elif s['dia'] is not None and s['hora']:
            # Coincidir día + hora + centro es "ya está"; sólo día + hora puede ser
            # otra cosa a la misma hora en otra sede, así que se avisa como posible.
            for dia, hora, tipo, centro, nota in existentes:
                if int(dia) != s['dia'] or hora != s['hora']:
                    continue
                mismo = s['lugar'] and centro and clave(s['lugar']) == clave(centro)
                etiqueta = f'{tipo} {hora} {centro}'.strip() + (f' «{nota}»' if nota else '')
                ya = ('sí — ' if mismo else 'quizá, misma hora en otra sede — ') + etiqueta
                if mismo:
                    break
        ws.append([DIAS[s['dia']] if s['dia'] is not None else '',
                   s['dia'] if s['dia'] is not None else '',
                   s['hora'] or '', s['hora_fin'] or '', s['titulo'], s['lugar'] or '',
                   s['pastoral'] or '', ya, s['original'], s['dudas']])
    anchos(ws, [12, 11, 8, 9, 44, 30, 13, 34, 46, 34])

    # --- Revisar ---
    ws = wb.create_sheet('Revisar')
    encabezar(ws, ['Qué', 'Fecha', 'Título', 'Celda origen', 'Duda', 'Texto original'])
    for a in cal.avisos:
        ws.append(['rejilla', '', '', '', a, ''])
    for e in eventos:
        if e['dudas'] or not e['inicio']:
            ws.append([e['clase'], e['inicio'].isoformat() if e['inicio'] else 'SIN FECHA',
                       e['titulo'], e['origen'],
                       e['dudas'] or 'sin fecha resuelta', e['original']])
    for s in semanales:
        if s['dudas']:
            ws.append(['semanal', '', s['titulo'], '', s['dudas'], s['original']])
    anchos(ws, [11, 12, 44, 16, 54, 46])

    # --- Resumen ---
    ws = wb.create_sheet('Resumen')
    ws.append(['Agenda 2026 — extracción automática'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(['Cuadros de texto en la hoja', len(hoja.cuadros)])
    ws.append(['Paneles «Actividades semanales» encontrados', n_paneles])
    ws.append(['Filas de evento generadas', len(eventos)])
    ws.append(['Filas de horario semanal', len(semanales)])
    ws.append(['Filas con dudas', sum(1 for e in eventos if e['dudas'] or not e['inicio'])])
    ws.append([])
    ws.append(['Mes', 'Eventos', 'Periodos', 'Mensuales'])
    for m in range(1, 13):
        delmes = [e for e in eventos if e['inicio'] and e['inicio'].month == m]
        ws.append([NOMMES[m],
                   sum(1 for e in delmes if e['clase'] == 'evento'),
                   sum(1 for e in delmes if e['clase'] == 'periodo'),
                   sum(1 for e in delmes if e['clase'] == 'mensual')])
    ws.append(['SIN FECHA', sum(1 for e in eventos if not e['inicio']), '', ''])
    ws.append([])
    ws.append(['Estructura de la hoja: DOM-MIE (blanco) y JUE-SAB (amarillo) son meses '
               'distintos; mes amarillo = 12 - mes blanco'])
    ws.append(['Banda', 'Columnas', 'DOM-MIE', 'JUE-SAB'])
    for banda, f_dias in ((1, 3), (2, 11)):
        for b in BLOQUES:
            amarillo = []
            for i in (5, 6, 7):
                v = cal.colmes.get((banda, b + i))
                if v and v[0] not in amarillo:
                    amarillo.append(v[0])
            blanco = cal.dommie.get((banda, b))
            hay = any(hoja.relleno.get(f'{col_nombre(b + i)}{f_dias}') == AMARILLO
                      for i in (5, 6, 7))
            ws.append([banda, f'{col_nombre(b)}-{col_nombre(b + 7)}',
                       NOMMES.get(blanco, '—'),
                       ', '.join(NOMMES[x] for x in amarillo) + ('' if hay else ' (sin amarillo)')
                       if amarillo else '—'])
    anchos(ws, [42, 12, 12, 30])
    ws.column_dimensions['A'].width = 46

    wb.save(ruta_salida)


COLORES = {'evento': '#1e4d8b', 'mensual': '#1e4d8b', 'periodo': '#55a54a'}


def pastoral_id_de(nombre):
    if not nombre:
        return None
    for pid, n, _pistas in PASTORALES:
        if clave(n) == clave(str(nombre)):
            return pid
    return None


def a_json(ruta_xlsx: str, ruta_json: str):
    """Pasa el xlsx ya revisado a JSON, que es lo que lee importar_agenda.php.

    El PHP de este XAMPP no trae la extensión zip, así que no puede abrir un
    xlsx: el paso por JSON es lo que permite que se revise en Excel y se importe
    después con las clases del proyecto.
    """
    import json
    from openpyxl import load_workbook

    wb = load_workbook(ruta_xlsx, data_only=True)

    def filas(hoja):
        ws = wb[hoja]
        cabeceras = [c.value for c in next(ws.iter_rows(max_row=1))]
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if any(v not in (None, '') for v in fila):
                yield dict(zip(cabeceras, fila))

    def texto(v):
        v = '' if v is None else str(v).strip()
        return v or None

    def fecha(v):
        if v is None or v == '':
            return None
        if isinstance(v, (datetime.date, datetime.datetime)):
            return v.strftime('%Y-%m-%d')
        return str(v)[:10]

    def hora(v):
        if v is None or v == '':
            return None
        if isinstance(v, datetime.time):
            return v.strftime('%H:%M')
        if isinstance(v, datetime.datetime):
            return v.strftime('%H:%M')
        return str(v)[:5]

    eventos, sin_fecha = [], 0
    for f in filas('Eventos'):
        inicio = fecha(f.get('Fecha inicio'))
        if not inicio:
            sin_fecha += 1
            continue
        clase = texto(f.get('Clasificación')) or 'evento'
        eventos.append(dict(
            fecha_inicio=inicio, fecha_fin=fecha(f.get('Fecha fin')),
            hora=hora(f.get('Hora')), hora_fin=hora(f.get('Hora fin')),
            titulo=texto(f.get('Título')), lugar=texto(f.get('Lugar')),
            pastoral_id=pastoral_id_de(f.get('Pastoral')), clase=clase,
            color=COLORES.get(clase, COLORES['evento']),
            origen=texto(f.get('Celda origen')),
            descripcion=texto(f.get('Texto original'))))

    horarios = []
    for f in filas('Horarios semanales'):
        dia = f.get('dia_semana')
        h = hora(f.get('Hora'))
        titulo = texto(f.get('Actividad'))
        if dia is None or dia == '' or not h or not titulo:
            continue
        if (texto(f.get('¿Ya existe?')) or '').lower().startswith('sí'):
            continue          # ya está cargado en horarios
        lugar = texto(f.get('Lugar'))
        horarios.append(dict(dia_semana=int(dia), hora=h, hora_fin=hora(f.get('Hora fin')),
                             titulo=titulo, lugar=lugar, centro_id=centro_id(lugar),
                             pastoral=texto(f.get('Pastoral'))))

    with io.open(ruta_json, 'w', encoding='utf-8') as f:
        json.dump(dict(origen=os.path.basename(ruta_xlsx), anio=ANIO,
                       eventos=eventos, horarios=horarios),
                  f, ensure_ascii=False, indent=1)
    print(f'Eventos ....................... {len(eventos)}'
          + (f'  ({sin_fecha} sin fecha, omitidos)' if sin_fecha else ''))
    print(f'Horarios semanales nuevos ..... {len(horarios)}')
    print(f'\nEscrito: {ruta_json}')


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('entrada', nargs='?', default='agenda 2026.xlsx')
    p.add_argument('-o', '--salida', default='agenda-2026-extraida.xlsx')
    p.add_argument('--a-json', metavar='ARCHIVO.json',
                   help='no extrae: convierte el xlsx ya revisado al JSON que '
                        'lee importar_agenda.php')
    args = p.parse_args()

    if args.a_json:
        revisado = args.entrada if args.entrada.endswith('.xlsx') and os.path.isfile(
            args.entrada) and args.entrada != 'agenda 2026.xlsx' else args.salida
        if not os.path.isfile(revisado):
            sys.exit(f'No encuentro «{revisado}».')
        print(f'Leyendo el xlsx revisado: {revisado}')
        a_json(revisado, args.a_json)
        return

    if not os.path.isfile(args.entrada):
        sys.exit(f'No encuentro «{args.entrada}».')

    hoja, cal, eventos, semanales, n_paneles = extraer(args.entrada)
    escribir(args.salida, hoja, cal, eventos, semanales, n_paneles)

    con_fecha = [e for e in eventos if e['inicio']]
    print(f'Cuadros en la hoja ............ {len(hoja.cuadros)}')
    print(f'Paneles semanales ............. {n_paneles} (se usa uno)')
    print(f'Filas de evento ............... {len(eventos)}  '
          f'({len(eventos) - len(con_fecha)} sin fecha)')
    print(f'  eventos ..................... {sum(1 for e in eventos if e["clase"] == "evento")}')
    print(f'  periodos (barras) ........... {sum(1 for e in eventos if e["clase"] == "periodo")}')
    print(f'  mensuales recurrentes ....... {sum(1 for e in eventos if e["clase"] == "mensual")}')
    print(f'Horarios semanales ............ {len(semanales)}')
    print(f'Filas con dudas ............... '
          f'{sum(1 for e in eventos if e["dudas"] or not e["inicio"])}')
    if cal.avisos:
        print(f'Avisos de la rejilla .......... {len(cal.avisos)}')
        for a in cal.avisos:
            print(f'  - {a}')
    pormes = Counter(e['inicio'].month for e in con_fecha)
    print('Por mes: ' + '  '.join(f'{NOMMES[m][:3]} {pormes.get(m, 0)}' for m in range(1, 13)))
    print(f'\nEscrito: {args.salida}')


if __name__ == '__main__':
    main()
